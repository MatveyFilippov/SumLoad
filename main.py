import shutil
import sys
import logging
import os
import pandas as pd
import numpy as np
import json
import pkg_resources
import openpyxl
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QPushButton,
    QFileDialog, QVBoxLayout, QHBoxLayout, QLineEdit, QComboBox
)

log_file_path = pkg_resources.resource_filename(__name__, 'SumLoadErrors.log')
json_set_file_path = pkg_resources.resource_filename(__name__, 'SumLoadDefaultSettings.json')
logging.basicConfig(filename=log_file_path, level=logging.ERROR, encoding="UTF-8", datefmt="%Y-%m-%d %H:%M:%S",
                    format="\n\n%(levelname)s %(message)s %(asctime)s")


def get_settings(name_of_param: str):
    try:
        with open(json_set_file_path, "r", encoding="UTF-8") as json_file:
            return json.load(json_file)[name_of_param]
    except Exception:
        return None


def get_sum(sheet_name: str, file_path: str, proc: str, thick: float, width: float, length: float):
    load = pd.read_excel(file_path, sheet_name=sheet_name)
    load = load.loc[:, ["Proc.", "Thick.", "Width", "Length", "Eff.sqm", "DONE_TAG"]]
    filtered = load[
        (load['Proc.'] == proc) &
        (load['Thick.'] == thick) &
        (load['Width'] == width) &
        (load["Length"] == length)
    ]

    try:
        tags = filtered[(filtered["DONE_TAG"] == "X")]

        # Определенный индекс, после которого нужно получить сумму
        start_index = list(tags.index)[len(list(tags.index)) - 1]  # Индекс последнего 'X' в колонке DONE_TAG
        index_in_filtered = filtered.index.get_loc(start_index)

        filtered = filtered.iloc[index_in_filtered:]
    except IndexError:
        raise Exception("НЕТОЧНЫЕ ВЫЧИСЛЕНИЯ: Не нашёл 'X' в столбце 'DONE_TAG'")

    return filtered["Eff.sqm"].sum()


class MainWindow(QMainWindow):

    def __init__(self):
        super().__init__()

        # Создаем виджеты
        self.get_log_btn = QPushButton('GetLog')
        self.get_log_btn.setMaximumSize(70, 30)

        self.open_button = QPushButton('Открыть файл')
        default_path = get_settings("file_path")
        if default_path and not os.path.exists(default_path):
            default_path = None
        self.file_path = f"{default_path if default_path is not None else 'Файл не выбран'}"
        self.file_label = QLabel(self.file_path)

        self.choice_sheet = QComboBox()
        self.choice_sheet.addItems(['Выберите название листа:'])
        self.choice_sheet.setDisabled(True)
        self.name_of_page = None
        self.find_button = QPushButton('Найти шаблоны')
        self.find_button.setMaximumSize(300, 50)
        self.find_button.setDisabled(True)
        self.choice_proc = QComboBox()
        self.choice_proc.addItems(['Выберите proc:'])
        self.choice_proc.setDisabled(True)
        self.choice_width = QComboBox()
        self.choice_width.addItems(['Выберите Width:'])
        self.choice_width.setDisabled(True)
        self.choice_thick = QComboBox()
        self.choice_thick.addItems(['Выберите Thick:'])
        self.choice_thick.setDisabled(True)
        self.choice_length = QComboBox()
        self.choice_length.addItems(['Выберите Length:'])
        self.choice_length.setDisabled(True)
        self.go_button = QPushButton('Суммировать')
        self.go_button.setDisabled(True)
        self.result_preview = QLabel("Здесь будет сумма по вашим шаблонам:")
        self.result_text = QLineEdit('')
        self.result_text.setReadOnly(True)

        if self.file_path != 'Файл не выбран':
            self.set_sheet_name()

        # Устанавливаем расположение виджетов
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout()
        central_widget.setLayout(layout)

        file_layout = QHBoxLayout()
        file_layout.addWidget(self.open_button)
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(self.get_log_btn)
        layout.addLayout(file_layout)

        layout.addWidget(self.choice_sheet)

        find_layout = QHBoxLayout()
        find_layout.addWidget(self.find_button)
        layout.addLayout(find_layout)

        layout.addWidget(self.choice_proc)
        layout.addWidget(self.choice_thick)
        layout.addWidget(self.choice_length)
        layout.addWidget(self.choice_width)

        result_layout = QHBoxLayout()
        result_layout.addWidget(self.go_button)
        result_layout.addWidget(self.result_preview)
        result_layout.addWidget(self.result_text)
        layout.addLayout(result_layout)

        # Настраиваем главное окно
        self.setWindowTitle('Получение суммы по шаблонам из Excel')
        self.setGeometry(200, 200, 700, 300)

        # Соединяем кнопки с функцией
        self.get_log_btn.clicked.connect(self.load_log)
        self.open_button.clicked.connect(self.open_file)
        self.go_button.clicked.connect(self.go)
        self.find_button.clicked.connect(self.find_pattern)

    def load_log(self):
        folder_path = QFileDialog.getExistingDirectory(self, 'Выберите папку куда сохранить SumLoadErrors.log')
        if not folder_path:
            self.print_("Не выбрана папка для сохранения файла с логами", red=True)
            return
        try:
            shutil.copy(log_file_path, folder_path)
        except Exception:
            raise Exception("Не удалось сохранить файл с логами")
        self.print_(f"{os.path.join(folder_path, 'SumLoadErrors.log')}", preview_text="Файл сохранён:")

    def print_(self, text: str, red=False, preview_text="Приложение:"):
        self.result_text.setText(text)
        if red:
            preview_text = "Внимание:"
            self.result_text.setStyleSheet("color: red;")
        else:
            self.result_text.setStyleSheet("color: black;")
        self.result_preview.setText(preview_text)

    def get_unique_values(self, load: pd.DataFrame, col_name) -> np.ndarray:
        load = load[col_name]

        # Парсинг уникальных значений
        unique_values = load.drop_duplicates().to_numpy()
        # Преобразуем всё в строку
        unique_values = unique_values.astype(str)
        # Создаем список слов, которые нужно исключить
        exclude_words = ['nun', '4']
        # Фильтруем массив, оставляя только элементы, не содержащие исключаемые слова и символы
        unique_values = np.extract(~np.isin(unique_values, exclude_words), unique_values)

        if col_name == "Thick.":
            mask = np.char.find(unique_values, ',') == -1
            if False in mask:
                raise Exception("В excel файле в столбце 'Thick.' стоят ',' -> нужно сделать '.'")

        try:
            unique_values = unique_values.astype(float)
            if len(set(list(unique_values))) != len(list(unique_values)):
                self.print_(
                    "НЕТОЧНОЫЕ ВЫЧИСЛЕНИЯ: Скорее всего в столбце 'Thick.' есть что-то по типу 2.6 & 2.60",
                    True
                )
            # Определяем порядок сортировки
            sort_order = np.argsort(unique_values)
            # Применяем порядок сортировки к исходному массиву
            unique_values = unique_values[sort_order]
            # Преобразуем всё в строку
            unique_values = unique_values.astype(str)
        except ValueError:
            unique_values = unique_values.astype(str)

        return unique_values

    def disable_all_buttons(self):
        self.choice_sheet.setDisabled(True)
        self.find_button.setDisabled(True)
        self.choice_proc.setDisabled(True)
        self.choice_width.setDisabled(True)
        self.choice_thick.setDisabled(True)
        self.choice_length.setDisabled(True)
        self.go_button.setDisabled(True)

    def open_file(self):
        self.disable_all_buttons()
        self.print_("", preview_text="Здесь будет сумма по вашим шаблонам:")

        # Открываем диалог выбора файла
        file_path, _ = QFileDialog.getOpenFileName(self, 'Выберите файл', '', 'Excel Files (*.xlsx *.xls)')

        # Если файл выбран
        if file_path:
            self.file_path = file_path
            # Обновляем текст метки
            self.file_label.setText(file_path)
            self.set_sheet_name()
        else:
            self.file_label.setText('Файл не выбран')
            self.print_('Заново выберите таблицу', True)

    def set_sheet_name(self):
        first_item = self.choice_sheet.itemText(0)
        self.choice_sheet.clear()
        self.choice_sheet.addItems([first_item])
        workbook = openpyxl.load_workbook(self.file_path)
        self.choice_sheet.addItems(workbook.sheetnames)
        try:
            def_num_thick = int(list(workbook.sheetnames).index(get_settings('sheet_name'))) + 1
            self.choice_sheet.setCurrentIndex(def_num_thick)
        except ValueError:
            pass
        self.choice_sheet.setDisabled(False)
        self.find_button.setDisabled(False)

    def put_params_in_btn(self, label: str, box: QComboBox, name_of_set: str, xlsx_list: pd.DataFrame):
        box_first_item = box.itemText(0)
        box.clear()
        box.addItems([box_first_item])
        try:
            params = self.get_unique_values(xlsx_list, label)
        except (ValueError, KeyError):
            raise Exception(f"Отсутствует столбец '{label}' на выбранном листе")
        box.addItems(params)
        try:
            def_num = int(list(params).index(get_settings(name_of_set))) + 1
            box.setCurrentIndex(def_num)
        except ValueError:
            pass
        box.setDisabled(False)

    def find_pattern(self):
        self.print_("", preview_text="Здесь будет сумма по вашим шаблонам:")
        self.name_of_page = self.choice_sheet.currentText()

        load = pd.read_excel(self.file_path, sheet_name=self.name_of_page)

        try:
            load["DONE_TAG"]
        except KeyError:
            raise Exception("Отсутствует столбец 'DONE_TAG' на выбранном листе")

        self.put_params_in_btn("Proc.", self.choice_proc, 'proc', load)
        self.put_params_in_btn("Thick.", self.choice_thick, 'thick', load)
        self.put_params_in_btn("Length", self.choice_length, 'length', load)
        self.put_params_in_btn("Width", self.choice_width, 'width', load)

        self.go_button.setDisabled(False)

    def go(self):
        data = {
            'sheet_name': self.name_of_page,
            'file_path': self.file_path,
            'proc': self.choice_proc.currentText(),
            'thick': self.choice_thick.currentText(),
            'length': self.choice_length.currentText(),
            'width': self.choice_width.currentText()
        }
        with open(json_set_file_path, "w", encoding="UTF-8") as json_file:
            json.dump(data, json_file)
        try:
            thick = float(self.choice_thick.currentText())
            width = float(self.choice_width.currentText())
            length = float(self.choice_length.currentText())
        except ValueError:
            self.print_("Ничего не найдено по этим шаблонам", red=True)
            return

        answer = get_sum(sheet_name=self.name_of_page, file_path=self.file_path,
                         proc=self.choice_proc.currentText(), thick=thick, length=length, width=width)
        # Выводим результат в окно
        if answer == 0:
            self.print_("Ничего не найдено по этим шаблонам", red=True)
        else:
            self.print_(str(answer), preview_text="Сумма по вашим параметрам:")


class ShowMustGoOn:
    def __init__(self, line: QLineEdit, preview: QLabel):
        self.line = line
        self.preview = preview

    def catcher(self, er_type, value, traceback):
        # Запись непредвиденных ошибок в лог файл
        logging.error("from", exc_info=(er_type, value, traceback))
        self.preview.setText("ОШИБКА:")
        self.line.setText(str(value))
        self.line.setStyleSheet("color: red;")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    sys.excepthook = ShowMustGoOn(window.result_text, window.result_preview).catcher
    window.show()
    sys.exit(app.exec_())
